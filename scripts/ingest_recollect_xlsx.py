#!/usr/bin/env python3
"""将清理后的重新采集xlsx导入知识库CSV。"""
import csv
import hashlib
import re
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
KB_FILE = REPO_ROOT / "knowledge_base_curated.csv"
XLSX_FILE = REPO_ROOT / ".tmp_recollect_final.xlsx"
TODAY = datetime.now().strftime("%Y-%m-%d")

HEADERS = [
    "id", "title", "category", "subcategory", "lab_type", "risk_level",
    "hazard_types", "scenario", "question", "answer", "steps", "ppe",
    "forbidden", "disposal", "first_aid", "emergency", "legal_notes",
    "references", "source_type", "source_title", "source_org",
    "source_version", "source_date", "source_url", "last_updated",
    "reviewer", "status", "tags", "language",
]

CATEGORY_MAP = {
    "大型分析仪器": "设备安全",
    "辐射安全": "辐射",
    "激光安全": "物理",
    "电气安全": "电气",
    "机械安全": "物理",
    "化学品SDS": "化学",
    "特定危害类型": "化学",
    "危废处置": "废弃物",
    "PPE专题深化": "通用",
    "实验室场景化应急": "通用",
    "培训体系": "培训",
    "管理制度": "通用",
    "通用安全": "通用",
    "标准法规": "标准",
    "高校手册": "通用",
    "绿色化学": "化学",
    "通用指南": "通用",
}

LAB_TYPE_MAP = {
    "设备安全": "通用", "辐射": "物理", "物理": "物理", "电气": "电气",
    "废弃物": "化学", "化学": "化学", "通用": "通用", "培训": "通用",
    "标准": "通用", "生物": "生物",
}

RISK_KEYWORDS = {
    5: ["爆炸", "死亡", "致命", "剧毒", "氰化物", "自燃", "IDLH", "立即危及生命"],
    4: ["火灾", "高压", "触电", "放射性", "辐射", "激光", "腐蚀", "强酸", "强碱", "强氧化", "强还原", "高温", "窒息"],
    3: ["易燃", "有毒", "有害气体", "泄漏", "烫伤", "灼伤", "割伤", "刺伤", "感染", "生物污染"],
    2: ["刺激性", "挥发性", "噪音", "粉尘"],
    1: ["标识", "培训", "检查", "记录", "清洁", "整理"],
}

ID_SEQ = 0


def next_id():
    global ID_SEQ
    ID_SEQ += 1
    return f"KB-RECO-{ID_SEQ:04d}"


def sig(title, question):
    return hashlib.md5(f"{title}||{question}".encode()).hexdigest()[:10]


def infer_risk_level(text):
    text = text.lower()
    for level, keywords in sorted(RISK_KEYWORDS.items(), reverse=True):
        for kw in keywords:
            if kw in text:
                return str(level)
    return "3"


def generate_question(title, category, subcategory):
    title = str(title).strip()
    cat = str(category).strip()
    sub = str(subcategory).strip()
    if "应急" in cat or "应急" in sub:
        return f"{title}应如何应急处置？"
    if "PPE" in cat or "手套" in title or "护目镜" in title:
        return f"{title}有哪些要求和注意事项？"
    if "培训" in cat or "制度" in cat:
        return f"关于{title}，有哪些规定和要求？"
    if "标准" in cat or "法规" in cat:
        return f"{title}的主要内容是什么？"
    if "废" in cat or "处置" in title:
        return f"{title}应如何安全处理？"
    if "辐射" in cat or "放射性" in title:
        return f"{title}的防护要求是什么？"
    if "激光" in cat or "激光" in title:
        return f"{title}的安全要求有哪些？"
    if "电气" in cat:
        return f"{title}应如何安全操作？"
    if "设备" in cat or "仪器" in cat:
        return f"{title}的安全操作规程是什么？"
    return f"关于{title}，有哪些安全要求和注意事项？"


def is_duplicate(entry, existing_titles, existing_sigs):
    title = str(entry.get("title", "")).strip()
    if title in existing_titles:
        return True, "exact_title"
    for et in existing_titles:
        if et and title and (et in title or title in et) and et != title:
            return True, f"fuzzy_match({et})"
    return False, ""


def main():
    import openpyxl
    if not XLSX_FILE.exists():
        print(f"ERROR: {XLSX_FILE} not found")
        return 1

    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb["安全数据总表"]

    # 读取现有知识库
    existing_rows = []
    existing_titles = set()
    existing_sigs = set()
    max_reco_num = 0

    if KB_FILE.exists():
        with KB_FILE.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                existing_rows.append(row)
                t = row.get("title", "").strip()
                if t:
                    existing_titles.add(t)
                existing_sigs.add(sig(t, row.get("question", "").strip()))
                rid = row.get("id", "")
                m = re.match(r"KB-RECO-(\d+)", rid)
                if m:
                    max_reco_num = max(max_reco_num, int(m.group(1)))

    print(f"现有知识库条目: {len(existing_rows)}")
    print(f"现有最大 KB-RECO ID: {max_reco_num}")

    global ID_SEQ
    ID_SEQ = max_reco_num

    skipped = {"exact_title": 0, "fuzzy_match": 0}
    new_rows = []
    dup_details = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = str(row[0]).strip() if row[0] else ""
        raw_cat = str(row[1]).strip() if row[1] else ""
        subcategory = str(row[2]).strip() if row[2] else ""
        batch = str(row[3]).strip() if row[3] else ""
        source_org = str(row[4]).strip() if row[4] else ""
        source_title = str(row[5]).strip() if row[5] else ""
        source_url = str(row[6]).strip() if row[6] else ""
        standard_code = str(row[7]).strip() if row[7] else ""
        clause_ref = str(row[8]).strip() if row[8] else ""
        hazards = str(row[9]).strip() if row[9] else ""
        safety_measures = str(row[10]).strip() if row[10] else ""
        emergency = str(row[11]).strip() if row[11] else ""
        ppe = str(row[12]).strip() if row[12] else ""
        procedures = str(row[13]).strip() if row[13] else ""
        legal_req = str(row[14]).strip() if row[14] else ""
        description = str(row[15]).strip() if row[15] else ""

        if not title:
            continue

        entry = {
            "title": title,
            "category": raw_cat,
            "subcategory": subcategory,
        }

        is_dup, reason = is_duplicate(entry, existing_titles, existing_sigs)
        if is_dup:
            if reason.startswith("fuzzy"):
                skipped["fuzzy_match"] += 1
            else:
                skipped["exact_title"] += 1
            dup_details.append(f"跳过: {title} | 原因: {reason}")
            continue

        category = CATEGORY_MAP.get(raw_cat, raw_cat)
        lab_type = LAB_TYPE_MAP.get(category, "通用")

        risk_text = f"{hazards} {safety_measures} {emergency} {title}"
        risk_level = infer_risk_level(risk_text)

        scenario = f"安全操作-{title}"
        if "应急" in raw_cat:
            scenario = f"应急响应-{title}"
        elif "标准" in raw_cat:
            scenario = f"标准法规-{title}"
        elif "设备" in raw_cat or "仪器" in raw_cat:
            scenario = f"设备操作-{title}"

        question = generate_question(title, category, subcategory)

        # 构建answer
        answer_parts = []
        if description:
            answer_parts.append(description)
        if hazards:
            answer_parts.append(f"主要危害：{hazards}")
        if safety_measures:
            answer_parts.append(f"安全措施：{safety_measures}")
        if legal_req:
            answer_parts.append(f"法规要求：{legal_req}")
        if emergency:
            answer_parts.append(f"应急处置：{emergency}")
        answer = "\n".join(answer_parts) if answer_parts else "请参照相关安全规范执行。"

        # source_type
        source_type = "authoritative_manual"
        if "OSHA" in source_org or "osha.gov" in source_url:
            source_type = "regulatory_standard"
        elif "CDC" in source_org or "cdc.gov" in source_url:
            source_type = "public_authoritative_source"
        elif "NRC" in source_org or "nrc.gov" in source_url:
            source_type = "public_authoritative_source"
        elif "EPA" in source_org or "epa.gov" in source_url:
            source_type = "public_authoritative_source"
        elif "标准" in raw_cat or "GB" in standard_code:
            source_type = "regulatory_standard"
        elif "高校" in raw_cat or ".edu" in source_url:
            source_type = "university_manual"

        # references
        references = source_title
        if standard_code:
            references = f"{standard_code} {source_title}"
        if clause_ref:
            references += f"; {clause_ref}"

        # legal_notes
        legal_notes = "遵守相关法律法规和实验室安全管理制度"
        if standard_code:
            legal_notes += f"；参照{standard_code}"

        # tags
        tags = f"重新采集;{batch}"
        if subcategory:
            tags += f";{subcategory}"

        kb_row = {
            "id": next_id(),
            "title": title,
            "category": category,
            "subcategory": subcategory,
            "lab_type": lab_type,
            "risk_level": risk_level,
            "hazard_types": hazards,
            "scenario": scenario,
            "question": question,
            "answer": answer,
            "steps": procedures,
            "ppe": ppe,
            "forbidden": "禁止未培训上岗；禁止违反操作规程；禁止擅自移除安全装置",
            "disposal": "按实验室废弃物分类管理制度收集处置",
            "first_aid": "发生伤害立即停止操作，按具体情况采取急救措施并就医",
            "emergency": emergency,
            "legal_notes": legal_notes,
            "references": references,
            "source_type": source_type,
            "source_title": source_title,
            "source_org": source_org,
            "source_version": standard_code,
            "source_date": "",
            "source_url": source_url,
            "last_updated": TODAY,
            "reviewer": "auto-ingest-recollect; pending human review",
            "status": "draft",
            "tags": tags,
            "language": "zh-CN",
        }

        s = sig(kb_row["title"], kb_row["question"])
        if s in existing_sigs:
            skipped["fuzzy_match"] += 1
            dup_details.append(f"跳过(签名重复): {title}")
            continue

        existing_sigs.add(s)
        existing_titles.add(title)
        new_rows.append(kb_row)

    print(f"\nXLSX读取记录: {ws.max_row - 1}")
    print(f"去重结果:")
    print(f"  完全重复title跳过: {skipped['exact_title']}")
    print(f"  模糊匹配跳过: {skipped['fuzzy_match']}")
    print(f"  实际新增: {len(new_rows)}")

    if not new_rows:
        print("没有新条目需要添加。")
        return 0

    all_rows = existing_rows + new_rows
    with KB_FILE.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        for row in all_rows:
            clean = {h: row.get(h, "") for h in HEADERS}
            writer.writerow(clean)

    print(f"\n完成。知识库总条目: {len(all_rows)}")

    # 生成报告
    report_path = REPO_ROOT / "recollect_ingest_report.md"
    batch_stats = {}
    cat_stats = {}
    for row in new_rows:
        b = row.get("tags", "").split(";")[0]
        batch_stats[b] = batch_stats.get(b, 0) + 1
        c = row.get("category", "")
        cat_stats[c] = cat_stats.get(c, 0) + 1

    with report_path.open("w", encoding="utf-8") as f:
        f.write("# 重新采集数据导入报告\n\n")
        f.write(f"**导入时间**: {TODAY}\n\n")
        f.write("## 统计摘要\n\n")
        f.write(f"| 指标 | 数值 |\n")
        f.write(f"|------|------|\n")
        f.write(f"| 导入前知识库条目 | {len(existing_rows)} |\n")
        f.write(f"| XLSX记录数 | {ws.max_row - 1} |\n")
        f.write(f"| 完全重复跳过 | {skipped['exact_title']} |\n")
        f.write(f"| 模糊匹配跳过 | {skipped['fuzzy_match']} |\n")
        f.write(f"| **实际新增** | **{len(new_rows)}** |\n")
        f.write(f"| 导入后知识库条目 | {len(all_rows)} |\n\n")

        f.write("## 按批次分布（新增）\n\n")
        f.write("| 批次 | 数量 |\n|------|------|\n")
        for k, v in sorted(batch_stats.items(), key=lambda x: -x[1]):
            f.write(f"| {k} | {v} |\n")
        f.write("\n")

        f.write("## 按分类分布（新增）\n\n")
        f.write("| 分类 | 数量 |\n|------|------|\n")
        for k, v in sorted(cat_stats.items(), key=lambda x: -x[1]):
            f.write(f"| {k} | {v} |\n")
        f.write("\n")

        f.write("## 去重详情\n\n")
        f.write("```\n")
        for d in dup_details:
            f.write(d + "\n")
        f.write("```\n")

    print(f"报告已生成: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
